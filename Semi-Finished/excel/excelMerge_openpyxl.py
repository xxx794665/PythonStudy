import xlrd
import openpyxl

source_file = []
for i in range(1, 269):
    file_name = "F:/xls/" + str(i) + "_body.xls"
    source_file.append(file_name)

target_file = "F:/xls/test.xls"    

data = []
for i in source_file:
    wb = xlrd.open_workbook(i, formatting_info=True)  # 保留原格式
    for sheet in wb.sheets():
        for rownum in range(sheet.nrows):
            data.append(sheet.row_values(rownum))
print(data)


workbook = openpyxl.Workbook(target_file)
worksheet = workbook.create_sheet()
rownum = 1
for i in range(len(data)):
    for j in range(len(data[i])):
        worksheet.cell(rownum, data[i][j])
        rownum += 1
workbook.save(target_file)
workbook.close()
